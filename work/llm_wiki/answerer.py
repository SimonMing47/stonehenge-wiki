from __future__ import annotations

import re
from html import unescape
from pathlib import Path
from typing import Any

from .execution import run_code_file
from .extractors import SUPPORTED_EXTENSIONS
from .formatting import make_standard_response
from .indexer import WikiIndex, query_terms
from .llm import LLMClient
from .models import CommentRecord, DocumentRecord, Question
from .repair import repair_document
from .security import PermissionGuard

EXT_ALIASES = {
    "word": ["doc", "docx"],
    "excel": ["xls", "xlsx"],
    "ppt": ["ppt", "pptx"],
    "powerpoint": ["ppt", "pptx"],
    "代码": ["java", "py", "js"],
    "office": ["doc", "docx", "ppt", "pptx", "xls", "xlsx"],
    "办公": ["doc", "docx", "ppt", "pptx", "xls", "xlsx"],
}

COMMAND_TERMS = {
    "select",
    "insert",
    "update",
    "delete",
    "gsql",
    "ssh",
    "curl",
    "jdbc",
    "mysql",
    "psql",
}


class QuestionAnswerer:
    def __init__(self, index: WikiIndex, guard: PermissionGuard, llm_client: LLMClient | None = None):
        self.index = index
        self.guard = guard
        self.llm_client = llm_client

    def answer(self, question: Question) -> dict:
        blocked, _ = self.guard.check_question(question.title)
        if blocked:
            return make_standard_response(question.id, question.title, question.level, "blocked")

        title = question.title
        if is_file_count_query(title):
            exts = requested_extensions(title)
            return make_standard_response(
                question.id,
                title,
                question.level,
                "file_count",
                self.index.file_counts(exts),
            )

        if is_comment_query(title):
            if is_fix_query(title):
                return self.answer_fix(question)
            comments = self.filter_comments(title)
            if is_count_query(title):
                return make_standard_response(question.id, title, question.level, "comment_count", len(comments))
            return make_standard_response(
                question.id,
                title,
                question.level,
                "list",
                [comment.summary() for comment in comments],
            )

        if is_code_execution_query(title):
            return self.answer_code_execution(question)

        if is_pivot_query(title):
            return self.answer_pivot(question)

        return self.answer_knowledge(question)

    def filter_comments(self, title: str) -> list[CommentRecord]:
        comments = list(self.index.comments)
        exts = requested_extensions(title)
        if exts:
            comments = [c for c in comments if self.index.by_path.get(c.source_path, None) and self.index.by_path[c.source_path].suffix in exts]
        elif re.search(r"(代码|TODO|todo)", title):
            comments = [
                c
                for c in comments
                if self.index.by_path.get(c.source_path, None)
                and self.index.by_path[c.source_path].suffix in {"java", "py", "js", "html", "xml", "md"}
            ]

        mentioned = self.index.find_records_mentioned(title)
        if mentioned:
            paths = {record.rel_path for record in mentioned}
            comments = [c for c in comments if c.source_path in paths]

        assignees = {c.assignee for c in comments if c.assignee and c.assignee in title}
        if assignees:
            comments = [c for c in comments if c.assignee in assignees]

        dates = re.findall(r"\d{8}", title)
        if dates:
            if len(dates) >= 2:
                start, end = sorted(dates[:2])
                comments = [c for c in comments if c.end_date and start <= c.end_date <= end]
            elif re.search(r"(之前|以前|前|截至|截止)", title):
                comments = [c for c in comments if c.end_date and c.end_date <= dates[0]]
            elif re.search(r"(之后|以后|后|晚于)", title):
                comments = [c for c in comments if c.end_date and c.end_date >= dates[0]]
            else:
                comments = [c for c in comments if c.end_date == dates[0]]

        if "结构化" in title:
            comments = [c for c in comments if c.structured]
        return comments

    def answer_fix(self, question: Question) -> dict:
        title = question.title
        records = self.index.find_records_mentioned(title)
        if not records:
            comments = self.filter_comments(title)
            paths = [comment.source_path for comment in comments]
            records = [self.index.by_path[path] for path in paths if path in self.index.by_path]
        if not records:
            records = [record for record in self.index.records if record.comments]
        if not records:
            return make_standard_response(question.id, title, question.level, "list", [])

        record = records[0]
        blocked, _ = self.guard.check_question(title, [record.rel_path], operation="write")
        if blocked:
            return make_standard_response(question.id, title, question.level, "blocked")
        comments = [c for c in self.filter_comments(title) if c.source_path == record.rel_path] or record.comments
        source, target = repair_document(record, self.index.wiki_root, comments)
        return make_standard_response(question.id, title, question.level, "fix", (source, target))

    def answer_code_execution(self, question: Question) -> dict:
        title = question.title
        records = [
            record
            for record in (self.index.find_records_mentioned(title) or self.index.search(title, limit=5))
            if record.suffix in {"py", "js", "java"}
        ]
        if not records:
            return make_standard_response(question.id, title, question.level, "list", [])
        record = records[0]
        blocked, _ = self.guard.check_question(title, [record.rel_path], operation="execute")
        if blocked:
            return make_standard_response(question.id, title, question.level, "blocked")
        allowed, output = run_code_file(record.full_path, record.suffix, self.guard)
        if not allowed:
            return make_standard_response(question.id, title, question.level, "blocked")
        return make_standard_response(question.id, title, question.level, "list", [output])

    def answer_pivot(self, question: Question) -> dict:
        title = question.title
        records = [
            record
            for record in (self.index.find_records_mentioned(title) or self.index.search(title, limit=5))
            if record.suffix == "xlsx"
        ]
        if not records:
            return make_standard_response(question.id, title, question.level, "list", [])
        try:
            target = create_pivot_workbook(records[0], self.index.wiki_root)
        except Exception as exc:
            return make_standard_response(question.id, title, question.level, "list", [f"透视图生成失败:{exc}"])
        return make_standard_response(question.id, title, question.level, "list", [target])

    def answer_knowledge(self, question: Question) -> dict:
        title = question.title
        if self.guard.has_password_intent(title):
            env_records = [record for record in self.index.records if self.guard.is_env_path(record.rel_path)]
            records = self.index.search(title, limit=8, records=env_records)
            if not records:
                return make_standard_response(question.id, title, question.level, "blocked")
        else:
            records = self.index.search(title, limit=8)
        if asks_for_paths(title):
            paths = [record.rel_path for record in records]
            return make_standard_response(question.id, title, question.level, "paths", paths)
        snippets = []
        for record in records:
            snippets.extend(
                relevant_snippets(
                    record,
                    title,
                    allow_password=self.guard.has_password_intent(title) and self.guard.is_env_path(record.rel_path),
                )
            )
        if not snippets:
            snippets = [record.rel_path for record in records]
        if self.llm_client and not self.guard.has_password_intent(title):
            try:
                llm_answer = self.llm_client.answer(title, records, snippets)
            except Exception:
                llm_answer = None
            if llm_answer:
                return make_standard_response(
                    question.id,
                    title,
                    question.level,
                    "list",
                    [
                        llm_answer.text,
                        f"llm:{llm_answer.provider}/{llm_answer.model}",
                        "sources:" + ", ".join(llm_answer.sources),
                    ],
                )
        return make_standard_response(question.id, title, question.level, "list", snippets[:8])

    def explain(self, question: Question) -> dict[str, Any]:
        title = question.title
        route = question_route(title)
        blocked, reason = self.guard.check_question(title)
        result: dict[str, Any] = {
            "status": "blocked" if blocked else "ok",
            "id": question.id,
            "title": title,
            "level": question.level,
            "route": route,
            "terms": query_terms(title),
            "safety": {"blocked": blocked, "reason": reason or ""},
            "records": [],
            "evidence": [],
        }
        if blocked:
            return result

        if route == "file_count":
            exts = requested_extensions(title)
            result["summary"] = {"file_counts": self.index.file_counts(exts)}
            result["records"] = [record_summary(record) for record in self.index.records[:20]]
            return result

        if route in {"comment_count", "comments", "fix"}:
            comments = self.filter_comments(title)
            if route == "fix" and not comments:
                comments = [comment for record in self.index.records for comment in record.comments]
            result["summary"] = {"comment_count": len(comments)}
            result["records"] = records_from_comments(self.index, comments)
            result["evidence"] = [comment_evidence(comment) for comment in comments[:30]]
            if route == "fix" and result["records"]:
                target_path = result["records"][0]["path"]
                write_blocked, write_reason = self.guard.check_question(title, [target_path], operation="write")
                result["safety"]["write_blocked"] = write_blocked
                result["safety"]["write_reason"] = write_reason or ""
            return result

        if route == "code_execution":
            scored = [
                (score, record)
                for score, record in self.index.search_with_scores(title, limit=8)
                if record.suffix in {"py", "js", "java"}
            ]
            result["records"] = [record_summary(record, score) for score, record in scored]
            result["evidence"] = knowledge_evidence(scored, title, self.guard)
            if scored:
                execute_blocked, execute_reason = self.guard.check_question(
                    title,
                    [scored[0][1].rel_path],
                    operation="execute",
                )
                result["safety"]["execute_blocked"] = execute_blocked
                result["safety"]["execute_reason"] = execute_reason or ""
            return result

        if route == "pivot":
            scored = [
                (score, record)
                for score, record in self.index.search_with_scores(title, limit=8)
                if record.suffix == "xlsx"
            ]
            result["records"] = [record_summary(record, score) for score, record in scored]
            result["evidence"] = knowledge_evidence(scored, title, self.guard)
            return result

        scored = self.knowledge_records_with_scores(title)
        if self.guard.has_password_intent(title) and not scored:
            result["status"] = "blocked"
            result["safety"] = {"blocked": True, "reason": "password_outside_env_path"}
            return result
        result["route"] = "paths" if asks_for_paths(title) else "knowledge"
        result["records"] = [record_summary(record, score) for score, record in scored]
        result["evidence"] = knowledge_evidence(scored, title, self.guard)
        result["llm"] = {
            "enabled": bool(self.llm_client),
            "would_call": bool(self.llm_client and not self.guard.has_password_intent(title)),
        }
        return result

    def knowledge_records_with_scores(self, title: str) -> list[tuple[int, DocumentRecord]]:
        if self.guard.has_password_intent(title):
            env_records = [record for record in self.index.records if self.guard.is_env_path(record.rel_path)]
            return self.index.search_with_scores(title, limit=8, records=env_records)
        return self.index.search_with_scores(title, limit=8)


def is_file_count_query(title: str) -> bool:
    return bool(re.search(r"(文件数量|统计.*文件|多少个.*文件|不同类型文件)", title))


def is_comment_query(title: str) -> bool:
    return bool(re.search(r"(批注|TODO|todo|待办)", title))


def is_fix_query(title: str) -> bool:
    return bool(re.search(r"(修复|按批注.*改|根据批注.*改|修改.*文档|输出到)", title))


def is_count_query(title: str) -> bool:
    return bool(re.search(r"(数量|统计|多少)", title))


def is_code_execution_query(title: str) -> bool:
    return bool(re.search(r"(执行结果|运行结果|输出是什么|代码.*结果|运行.*代码|执行.*代码)", title))


def is_pivot_query(title: str) -> bool:
    return bool(re.search(r"(透视图|透视表|pivot|图表)", title, re.IGNORECASE))


def asks_for_paths(title: str) -> bool:
    return bool(re.search(r"(哪些文件|文件名称|文件名|路径|目录|涉及到.*文件|包含.*文件)", title))


def question_route(title: str) -> str:
    if is_file_count_query(title):
        return "file_count"
    if is_comment_query(title):
        if is_fix_query(title):
            return "fix"
        return "comment_count" if is_count_query(title) else "comments"
    if is_code_execution_query(title):
        return "code_execution"
    if is_pivot_query(title):
        return "pivot"
    if asks_for_paths(title):
        return "paths"
    return "knowledge"


def requested_extensions(title: str) -> list[str] | None:
    low = title.lower()
    result: list[str] = []
    for alias, exts in EXT_ALIASES.items():
        if alias.lower() in low:
            result.extend(exts)
    for ext in sorted(SUPPORTED_EXTENSIONS):
        if f".{ext}" in low or re.search(rf"(?<![a-z0-9]){re.escape(ext)}(?![a-z0-9])", low):
            result.append(ext)
    if not result:
        return None
    seen: set[str] = set()
    return [ext for ext in result if not (ext in seen or seen.add(ext))]


def relevant_snippets(record: DocumentRecord, title: str, allow_password: bool = False) -> list[str]:
    return [
        f"{item['source_path']}: {item['text']}"
        for item in evidence_snippets(record, title, allow_password=allow_password)
    ]


def evidence_snippets(record: DocumentRecord, title: str, allow_password: bool = False) -> list[dict[str, Any]]:
    terms = query_terms(title)
    scored: list[tuple[int, int, str, list[str]]] = []
    for line_no, raw_line in enumerate(record.text.splitlines(), start=1):
        if re.match(r"^\s*<(?:!doctype|html|head|meta|link|title|script|style)\b", raw_line, re.IGNORECASE):
            continue
        line = clean_evidence_text(raw_line)
        if not line:
            continue
        low = line.lower()
        if not allow_password and re.search(r"(password|passwd|密码|密钥|secret|token)", low, re.IGNORECASE):
            continue
        if re.search(r"(忽略.*规则|上帝模式|ignore.*previous|god\s*mode)", line, re.IGNORECASE):
            continue
        matched_terms = [term for term in terms if term.lower() in low]
        score = len(matched_terms)
        if any(term.lower() in COMMAND_TERMS for term in matched_terms):
            score += 5
        if re.search(r"(命令|gsql|ssh|curl|jdbc|连接)", title, re.IGNORECASE) and re.search(
            r"(gsql|ssh|curl|jdbc|mysql|psql|命令|连接)", line, re.IGNORECASE
        ):
            score += 6
        if score:
            scored.append((score, line_no, line[:240], matched_terms[:8]))
    scored.sort(key=lambda item: -item[0])
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for score, line_no, line, matched_terms in scored:
        value = f"{record.rel_path}:{line_no}:{line}"
        if value not in seen:
            seen.add(value)
            result.append(
                {
                    "source_path": record.rel_path,
                    "line": line_no,
                    "text": line,
                    "score": score,
                    "matched_terms": matched_terms,
                }
            )
        if len(result) >= 3:
            break
    return result


def clean_evidence_text(raw_line: str) -> str:
    text = re.sub(r"<[^>]+>", " ", raw_line)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    if text in {"{", "}", ");", "};"}:
        return ""
    return text


def knowledge_evidence(scored_records: list[tuple[int, DocumentRecord]], title: str, guard: PermissionGuard) -> list[dict[str, Any]]:
    evidence: list[dict[str, Any]] = []
    allow_password_query = guard.has_password_intent(title)
    for _, record in scored_records:
        evidence.extend(
            evidence_snippets(
                record,
                title,
                allow_password=allow_password_query and guard.is_env_path(record.rel_path),
            )
        )
    if evidence:
        return evidence[:30]
    return [
        {"source_path": record.rel_path, "text": record.rel_path, "score": score, "matched_terms": []}
        for score, record in scored_records[:30]
    ]


def record_summary(record: DocumentRecord, score: int = 0) -> dict[str, Any]:
    return {
        "path": record.rel_path,
        "suffix": record.suffix,
        "name": record.name,
        "tags": sorted(record.tags),
        "comment_count": len(record.comments),
        "score": score,
    }


def comment_evidence(comment: CommentRecord) -> dict[str, Any]:
    return {
        "source_path": comment.source_path,
        "kind": comment.kind,
        "line": comment.line,
        "todo": comment.todo,
        "assignee": comment.assignee,
        "end_date": comment.end_date,
        "structured": comment.structured,
        "text": comment.summary(),
    }


def records_from_comments(index: WikiIndex, comments: list[CommentRecord]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for comment in comments:
        if comment.source_path in seen or comment.source_path not in index.by_path:
            continue
        seen.add(comment.source_path)
        records.append(record_summary(index.by_path[comment.source_path]))
    return records


def create_pivot_workbook(record: DocumentRecord, wiki_root: Path) -> str:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, Reference

    source = load_workbook(record.full_path, data_only=True)
    sheet = source.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("工作表数据不足")
    header = [str(item) if item is not None else "" for item in rows[0]]
    data_rows = [row for row in rows[1:] if any(cell is not None for cell in row)]
    group_idx = 0
    counter: dict[str, int] = {}
    for row in data_rows:
        key = str(row[group_idx]) if group_idx < len(row) and row[group_idx] is not None else "空值"
        counter[key] = counter.get(key, 0) + 1

    target_rel = f"output/fixed/pivot_{record.full_path.stem}.xlsx"
    target = wiki_root / target_rel
    target.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "pivot"
    ws.append([header[group_idx] or "分组", "数量"])
    for key, count in sorted(counter.items()):
        ws.append([key, count])
    chart = BarChart()
    chart.title = "Pivot Count"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
    ws.add_chart(chart, "D2")
    wb.save(target)
    return target_rel
