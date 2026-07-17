from __future__ import annotations

import csv
import json
import os
import re
import tempfile
import zipfile
from html import unescape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from .execution import run_code_file
from .extractors import COUNT_EXTENSIONS, SUPPORTED_EXTENSIONS
from .formatting import make_standard_response
from .indexer import WikiIndex, query_terms
from .llm import LLMClient
from .models import CommentRecord, DocumentRecord, Question
from .office_bridge import LEGACY_TO_MODERN, convert_office
from .repair import extract_replacements, repair_document, repair_target_rel
from .security import PermissionGuard

EXT_ALIASES = {
    "word": ["doc", "docx"],
    "excel": ["xls", "xlsx"],
    "ppt": ["ppt", "pptx"],
    "powerpoint": ["ppt", "pptx"],
    "代码": ["java", "py", "js"],
    "office": ["doc", "docx", "ppt", "pptx", "xls", "xlsx"],
    "办公": ["doc", "docx", "ppt", "pptx", "xls", "xlsx"],
}

COMMAND_TERMS = {
    "select",
    "insert",
    "update",
    "delete",
    "gsql",
    "ssh",
    "curl",
    "jdbc",
    "mysql",
    "psql",
}


class QuestionAnswerer:
    def __init__(
        self,
        index: WikiIndex,
        guard: PermissionGuard,
        llm_client: LLMClient | None = None,
        llm_clients: dict[str, LLMClient] | None = None,
        default_agent: str = "default",
        source_agent_map: dict[str, str] | None = None,
    ):
        self.index = index
        self.guard = guard
        self.llm_client = llm_client
        self.llm_clients = llm_clients or {}
        self.default_agent = default_agent
        self.source_agent_map = source_agent_map or {}

    @staticmethod
    def source_category(rel_path: str) -> str:
        parts = rel_path.split("/")
        if len(parts) >= 3 and parts[0] == "docs":
            return parts[1]
        return "uncategorized"

    def _select_llm_client(self, records: list[DocumentRecord]) -> LLMClient | None:
        if not self.llm_clients:
            return self.llm_client
        if not records:
            return self.llm_clients.get(self.default_agent, self.llm_client)
        agent_names = {
            self.source_agent_map.get(self.source_category(record.rel_path), self.default_agent)
            for record in records
        }
        if len(agent_names) == 1:
            target = next(iter(agent_names))
            return self.llm_clients.get(target) or self.llm_clients.get(self.default_agent) or self.llm_client
        return self.llm_clients.get(self.default_agent, self.llm_client)

    def plan_questions(self, questions: list[Question]) -> dict[str, dict[str, Any]]:
        """Batch semantic routing/safety review through the restricted Agent."""

        client = self.llm_clients.get(self.default_agent) or self.llm_client
        judge = getattr(client, "judge_questions", None)
        if not callable(judge):
            return {}
        payload: list[dict[str, Any]] = []
        for question in questions:
            blocked, _ = self.guard.check_question(question.title)
            if blocked:
                # A known malicious title is never allowed to contaminate the
                # shared semantic batch. The normal answer path will reject it.
                continue
            source_risks: set[str] = set()
            for record in self.index.find_records_mentioned(question.title)[:8]:
                if self.guard.path_blocked(record.rel_path, operation="read"):
                    source_risks.add("permission_denied")
                if self.guard.contains_prompt_injection(record.rel_path) or self.guard.contains_prompt_injection(
                    record.text
                ):
                    source_risks.add("prompt_injection")
                if any(self.guard.contains_prompt_injection(comment.raw_text) for comment in record.comments):
                    source_risks.add("prompt_injection_comment")
            payload.append(
                {
                    "id": question.id,
                    "title": question.title,
                    "source_risks": sorted(source_risks),
                }
            )
        try:
            return judge(payload)
        except Exception:
            return {}

    def answer(self, question: Question, decision: dict[str, Any] | None = None) -> dict:
        blocked, _ = self.guard.check_question(question.title)
        if blocked:
            return make_standard_response(question.id, question.title, question.level, "blocked")

        title = question.title
        expected_kind = expected_answer_kind(question.answer_format)
        if expected_kind == "blocked":
            return make_standard_response(question.id, title, question.level, "blocked")
        if decision and decision.get("unsafe") is True and not self.guard.has_password_intent(title):
            return make_standard_response(question.id, title, question.level, "blocked")

        route = {
            "file_count": "file_count",
            "comment_count": "comment_count",
            "fix": "fix",
        }.get(expected_kind or "")
        fixed_route = question_route(title)
        if route is None and fixed_route != "knowledge":
            # Explicit competition contracts remain deterministic. The Agent
            # handles genuinely ambiguous semantics, but a model mistake cannot
            # turn e.g. a doc count into a knowledge answer.
            route = fixed_route
        if route is None and decision and decision.get("route") in AGENT_ROUTES:
            route = str(decision["route"])
        if route is None:
            route = fixed_route

        if route == "file_count":
            exts = answer_format_extensions(question.answer_format) or requested_count_extensions(title)
            return make_standard_response(
                question.id,
                title,
                question.level,
                "file_count",
                self.index.file_counts(exts),
            )

        if route == "code_execution":
            return self.answer_code_execution(question)

        if route == "pivot":
            return self.answer_pivot(question)

        if route in {"comment_count", "comments", "fix"}:
            if route == "fix":
                return self.answer_fix(question)
            comments = self.filter_comments(title)
            if route == "comment_count":
                return make_standard_response(question.id, title, question.level, "comment_count", len(comments))
            return make_standard_response(
                question.id,
                title,
                question.level,
                "list",
                [comment_answer_text(comment) for comment in comments],
            )

        return self.answer_knowledge(question, force_paths=route == "paths")

    def filter_comments(self, title: str) -> list[CommentRecord]:
        comments = list(self.index.comments)
        exts = requested_extensions(title)
        if exts:
            comments = [c for c in comments if self.index.by_path.get(c.source_path, None) and self.index.by_path[c.source_path].suffix in exts]
        elif re.search(r"代码", title):
            comments = [
                c
                for c in comments
                if self.index.by_path.get(c.source_path, None)
                and self.index.by_path[c.source_path].suffix in {"java", "py", "js", "html", "xml", "md"}
            ]

        mentioned = self.index.find_records_mentioned(title)
        if mentioned:
            paths = {record.rel_path for record in mentioned}
            comments = [c for c in comments if c.source_path in paths]

        assignees = requested_assignees(title, comments)
        if assignees:
            comments = [
                c
                for c in comments
                if any(
                    candidate and candidate.casefold() in {value.casefold() for value in assignees}
                    for candidate in (c.assignee, c.author)
                )
            ]

        dates = normalized_dates(title)
        date_filter_intent = bool(
            re.search(
                r"(end[\s_-]*date|结束日期|截止|截至|之前|以前|之后|以后|晚于|早于|到期|期限|"
                r"按(?:照)?日期|日期筛选|\d{8}\s*(?:的)?\s*(?:批注|TODO|待办))",
                title,
                re.IGNORECASE,
            )
        ) or len(dates) >= 2
        if dates and date_filter_intent:
            use_created = bool(
                re.search(
                    r"(?:创建|添加|发表).{0,5}(?:日期|时间)|(?:批注|评论).{0,3}(?:创建时间|添加时间|发表时间)",
                    title,
                )
            )

            def comment_date(comment: CommentRecord) -> str | None:
                if use_created and comment.created:
                    values = normalized_dates(comment.created)
                    return values[0] if values else None
                return comment.end_date

            if len(dates) >= 2:
                start, end = sorted(dates[:2])
                comments = [c for c in comments if comment_date(c) and start <= str(comment_date(c)) <= end]
            elif re.search(r"(之前|以前|前|截至|截止)", title):
                comments = [c for c in comments if comment_date(c) and str(comment_date(c)) <= dates[0]]
            elif re.search(r"(之后|以后|后|晚于)", title):
                comments = [c for c in comments if comment_date(c) and str(comment_date(c)) >= dates[0]]
            else:
                comments = [c for c in comments if comment_date(c) == dates[0]]

        if "结构化" in title:
            comments = [c for c in comments if c.structured]
        if re.search(r"(按|按照).{0,4}(日期|时间)|(日期|时间).{0,4}(排序|顺序)", title):
            comments.sort(key=lambda item: (item.end_date or "99999999", item.source_path, item.line or 0))
        elif re.search(r"(按|按照).{0,4}(责任人|负责人|处理人)", title):
            comments.sort(key=lambda item: (item.assignee or item.author or "", item.source_path, item.line or 0))
        return comments

    def answer_fix(self, question: Question) -> dict:
        title = question.title
        records = self.index.find_records_mentioned(title)
        if not records:
            comments = self.filter_comments(title)
            paths = [comment.source_path for comment in comments]
            records = [self.index.by_path[path] for path in paths if path in self.index.by_path]
        if not records:
            records = [record for record in self.index.records if record.comments]
        if not records:
            return make_standard_response(question.id, title, question.level, "list", [])

        record = records[0]
        if (
            self.guard.contains_prompt_injection(record.rel_path)
            or self.guard.contains_prompt_injection(record.text)
            or any(self.guard.contains_prompt_injection(comment.raw_text) for comment in record.comments)
        ):
            return make_standard_response(question.id, title, question.level, "blocked")
        try:
            target_rel = repair_target_rel(record.rel_path)
        except ValueError:
            return make_standard_response(question.id, title, question.level, "blocked")
        blocked, _ = self.guard.check_question(
            title, [record.rel_path, target_rel], operation="write"
        )
        if blocked:
            return make_standard_response(question.id, title, question.level, "blocked")
        comments = [c for c in self.filter_comments(title) if c.source_path == record.rel_path] or record.comments
        replacements = extract_replacements(comments)
        if not replacements:
            client = self._select_llm_client([record])
            propose = getattr(client, "propose_replacements", None)
            if callable(propose):
                try:
                    replacements = propose(title, record, comments)
                except Exception:
                    replacements = []
        source, target = repair_document(
            record,
            self.index.wiki_root,
            comments,
            replacements=replacements,
        )
        return make_standard_response(question.id, title, question.level, "fix", (source, target))

    def answer_code_execution(self, question: Question) -> dict:
        title = question.title
        candidates = [record for record in self.index.records if record.suffix in {"py", "js", "java"}]
        mentioned = [
            record for record in self.index.find_records_mentioned(title) if record.suffix in {"py", "js", "java"}
        ]
        records = mentioned or self.index.search(title, limit=5, records=candidates)
        if not records:
            return make_standard_response(question.id, title, question.level, "list", [])
        record = records[0]
        blocked, _ = self.guard.check_question(title, [record.rel_path], operation="execute")
        if blocked:
            return make_standard_response(question.id, title, question.level, "blocked")
        allowed, output = run_code_file(record.full_path, record.suffix, self.guard)
        if not allowed:
            return make_standard_response(question.id, title, question.level, "blocked")
        return make_standard_response(question.id, title, question.level, "list", [output])

    def answer_pivot(self, question: Question) -> dict:
        title = question.title
        candidates = [
            record
            for record in self.index.records
            if record.suffix in {"xls", "xlsx"}
            and not self.guard.path_blocked(record.rel_path, operation="read")
        ]
        mentioned = [
            record for record in self.index.find_records_mentioned(title) if record.suffix in {"xls", "xlsx"}
        ]
        records = mentioned or self.index.search(title, limit=5, records=candidates)
        if not records:
            return make_standard_response(question.id, title, question.level, "list", [])
        source_blocked, _ = self.guard.check_question(title, [records[0].rel_path], operation="read")
        if source_blocked:
            return make_standard_response(question.id, title, question.level, "blocked")
        pivot_targets = [
            f"output/fixed/pivot_{records[0].full_path.stem}.xlsx",
            f"output/fixed/pivot_{records[0].full_path.stem}.csv",
        ]
        if any(self.guard.path_blocked(path, operation="write") for path in pivot_targets):
            return make_standard_response(question.id, title, question.level, "blocked")
        try:
            target = create_pivot_workbook(records[0], self.index.wiki_root, title)
        except Exception as exc:
            return make_standard_response(question.id, title, question.level, "list", [f"透视图生成失败:{exc}"])
        return make_standard_response(question.id, title, question.level, "list", [target])

    def answer_knowledge(self, question: Question, force_paths: bool = False) -> dict:
        title = question.title
        path_intent = force_paths or asks_for_paths(title)
        search_limit = max(len(self.index.records), 8) if path_intent else 8
        mentioned = self.index.find_records_mentioned(title)
        if self.guard.has_password_intent(title):
            if mentioned and any(not self.guard.is_env_path(record.rel_path) for record in mentioned):
                return make_standard_response(question.id, title, question.level, "blocked")
            env_records = [record for record in self.index.records if self.guard.is_env_path(record.rel_path)]
            file_names = mentioned_file_names(title)
            env_names = {record.name.casefold() for record in env_records}
            if not mentioned and file_names and any(name.casefold() not in env_names for name in file_names):
                return make_standard_response(question.id, title, question.level, "blocked")
            records = [record for record in mentioned if self.guard.is_env_path(record.rel_path)] or self.index.search(
                title, limit=search_limit, records=env_records
            )
            if not records:
                return make_standard_response(question.id, title, question.level, "blocked")
        else:
            records = mentioned if path_intent and mentioned else self.index.search(title, limit=search_limit)
        if path_intent:
            paths = sorted({record.rel_path for record in records})
            return make_standard_response(question.id, title, question.level, "paths", paths)
        if self.guard.has_password_intent(title):
            blocked, _ = self.guard.check_question(title, [record.rel_path for record in records])
            if blocked:
                return make_standard_response(question.id, title, question.level, "blocked")
            values = credential_values(records, title)
            if values:
                return make_standard_response(question.id, title, question.level, "list", values)
        snippets = []
        for record in records:
            snippets.extend(
                relevant_snippets(
                    record,
                    title,
                    allow_password=self.guard.has_password_intent(title) and self.guard.is_env_path(record.rel_path),
                )
            )
        if not snippets:
            snippets = [record.rel_path for record in records]
        llm_client = self._select_llm_client(records)
        llm_safe = not any(
            self.guard.contains_prompt_injection(record.rel_path)
            or self.guard.contains_prompt_injection(record.text)
            for record in records
        )
        if llm_client and llm_safe and not self.guard.has_password_intent(title):
            try:
                llm_answer = llm_client.answer(title, records, snippets)
            except Exception:
                llm_answer = None
            if llm_answer:
                datas = llm_answer_datas(llm_answer.text)
                return make_standard_response(
                    question.id,
                    title,
                    question.level,
                    "list",
                    datas,
                )
        return make_standard_response(question.id, title, question.level, "list", snippets[:8])

    def explain(self, question: Question) -> dict[str, Any]:
        title = question.title
        route_hint = expected_answer_kind(question.answer_format)
        route = {
            "file_count": "file_count",
            "comment_count": "comment_count",
            "fix": "fix",
        }.get(route_hint or "", question_route(title))
        blocked, reason = self.guard.check_question(title)
        result: dict[str, Any] = {
            "status": "blocked" if blocked else "ok",
            "id": question.id,
            "title": title,
            "level": question.level,
            "route": route,
            "terms": query_terms(title),
            "safety": {"blocked": blocked, "reason": reason or ""},
            "records": [],
            "evidence": [],
        }
        if blocked:
            return result

        if route == "file_count":
            exts = answer_format_extensions(question.answer_format) or requested_count_extensions(title)
            result["summary"] = {"file_counts": self.index.file_counts(exts)}
            result["records"] = [record_summary(record) for record in self.index.records[:20]]
            return result

        if route in {"comment_count", "comments", "fix"}:
            comments = self.filter_comments(title)
            if route == "fix" and not comments:
                comments = [comment for record in self.index.records for comment in record.comments]
            result["summary"] = {"comment_count": len(comments)}
            result["records"] = records_from_comments(self.index, comments)
            result["evidence"] = [comment_evidence(comment) for comment in comments[:30]]
            if route == "fix" and result["records"]:
                target_path = result["records"][0]["path"]
                write_blocked, write_reason = self.guard.check_question(title, [target_path], operation="write")
                result["safety"]["write_blocked"] = write_blocked
                result["safety"]["write_reason"] = write_reason or ""
            return result

        if route == "code_execution":
            scored = [
                (score, record)
                for score, record in self.index.search_with_scores(title, limit=8)
                if record.suffix in {"py", "js", "java"}
            ]
            result["records"] = [record_summary(record, score) for score, record in scored]
            result["evidence"] = knowledge_evidence(scored, title, self.guard)
            if scored:
                execute_blocked, execute_reason = self.guard.check_question(
                    title,
                    [scored[0][1].rel_path],
                    operation="execute",
                )
                result["safety"]["execute_blocked"] = execute_blocked
                result["safety"]["execute_reason"] = execute_reason or ""
            return result

        if route == "pivot":
            scored = [
                (score, record)
                for score, record in self.index.search_with_scores(title, limit=8)
                if record.suffix in {"xls", "xlsx"}
            ]
            result["records"] = [record_summary(record, score) for score, record in scored]
            result["evidence"] = knowledge_evidence(scored, title, self.guard)
            return result

        scored = self.knowledge_records_with_scores(title)
        mentioned = self.index.find_records_mentioned(title)
        if self.guard.has_password_intent(title) and any(
            not self.guard.is_env_path(record.rel_path) for record in mentioned
        ):
            result["status"] = "blocked"
            result["safety"] = {"blocked": True, "reason": "password_outside_env_path"}
            return result
        if self.guard.has_password_intent(title) and not scored:
            result["status"] = "blocked"
            result["safety"] = {"blocked": True, "reason": "password_outside_env_path"}
            return result
        result["route"] = "paths" if asks_for_paths(title) else "knowledge"
        result["records"] = [record_summary(record, score) for score, record in scored]
        result["evidence"] = knowledge_evidence(scored, title, self.guard)
        result["llm"] = {
            "enabled": bool(self.llm_client),
            "would_call": bool(self.llm_client and not self.guard.has_password_intent(title)),
        }
        return result

    def knowledge_records_with_scores(self, title: str) -> list[tuple[int, DocumentRecord]]:
        if self.guard.has_password_intent(title):
            env_records = [record for record in self.index.records if self.guard.is_env_path(record.rel_path)]
            return self.index.search_with_scores(title, limit=8, records=env_records)
        return self.index.search_with_scores(title, limit=8)


def is_file_count_query(title: str) -> bool:
    if re.search(r"(文件.{0,3}(?:总数|数量|个数)|多少个.*文件|不同类型文件|各(?:类|种).*文件)", title):
        return True
    extension = r"(?:docx?|pptx?|xlsx?|xml|java|py|html|md|js)"
    return bool(
        re.search(rf"(?:统计|计算|查询).{{0,24}}(?<![A-Za-z0-9.]){extension}(?![A-Za-z0-9]).{{0,12}}(?:总数|数量|个数)", title, re.IGNORECASE)
        or re.search(rf"(?<![A-Za-z0-9.]){extension}(?![A-Za-z0-9]).{{0,12}}(?:总数|数量|个数)", title, re.IGNORECASE)
        or re.search(rf"(?<![A-Za-z0-9.]){extension}(?![A-Za-z0-9]).{{0,12}}(?:有|共|总共)?多少(?:个|份)?", title, re.IGNORECASE)
    )


def is_comment_query(title: str) -> bool:
    return bool(re.search(r"(批注|TODO|todo|待办)", title))


def is_fix_query(title: str) -> bool:
    return bool(
        re.search(
            r"(修复|按(?:照)?批注.*(?:改|处理|执行)|根据批注.*(?:改|处理|执行)|"
            r"修改.*文档|输出到|批注.{0,8}(?:优化|整理|落实|处理)|(?:优化|整理).{0,8}批注)",
            title,
        )
    )


def is_count_query(title: str) -> bool:
    return bool(re.search(r"(数量|总数|个数|多少|统计.{0,20}(?:数量|总数|个数))", title))


def is_list_query(title: str) -> bool:
    return bool(re.search(r"(列表|清单|明细|列出|有哪些|逐条)", title))


def is_code_execution_query(title: str) -> bool:
    return bool(
        re.search(r"(执行结果|运行结果|输出是什么|代码.*结果|运行.*代码|执行.*代码)", title)
        or re.search(
            r"(?:运行|执行).{0,80}(?:\.(?:py|js|java)|函数|方法|程序).{0,30}(?:结果|输出|得到|返回|是什么)",
            title,
            re.IGNORECASE,
        )
    )


def is_pivot_query(title: str) -> bool:
    return bool(re.search(r"(透视图|透视表|pivot|图表)", title, re.IGNORECASE))


def asks_for_paths(title: str) -> bool:
    return bool(re.search(r"(哪些文件|文件名称|文件名|路径|目录|涉及到.*文件|包含.*文件)", title))


def mentioned_file_names(title: str) -> set[str]:
    extensions = "|".join(sorted((re.escape(ext) for ext in SUPPORTED_EXTENSIONS), key=len, reverse=True))
    pattern = rf"[A-Za-z0-9_\u4e00-\u9fff-]+\.(?:{extensions})(?![A-Za-z0-9])"
    return {match.group(0) for match in re.finditer(pattern, title, re.IGNORECASE)}


def question_route(title: str) -> str:
    if is_file_count_query(title):
        return "file_count"
    if is_comment_query(title):
        if is_fix_query(title):
            return "fix"
        return "comment_count" if is_count_query(title) else "comments"
    if is_code_execution_query(title):
        return "code_execution"
    if is_pivot_query(title):
        return "pivot"
    if asks_for_paths(title):
        return "paths"
    return "knowledge"


AGENT_ROUTES = {
    "file_count",
    "comment_count",
    "comments",
    "fix",
    "code_execution",
    "pivot",
    "paths",
    "knowledge",
}


def requested_extensions(title: str) -> list[str] | None:
    low = title.lower()
    result: list[str] = []
    for ext in sorted(SUPPORTED_EXTENSIONS, key=lambda item: (-len(item), item)):
        if f".{ext}" in low or re.search(rf"(?<![a-z0-9]){re.escape(ext)}(?![a-z0-9])", low):
            result.append(ext)
    if not result:
        for alias, exts in EXT_ALIASES.items():
            alias_low = alias.lower()
            if re.search(rf"(?<![a-z0-9]){re.escape(alias_low)}(?![a-z0-9])", low):
                result.extend(exts)
    if not result:
        return None
    seen: set[str] = set()
    return [ext for ext in result if not (ext in seen or seen.add(ext))]


def requested_count_extensions(title: str) -> list[str] | None:
    """Return only the twelve file-count enum values allowed by the judge."""

    requested = requested_extensions(title) or []
    result = [extension for extension in requested if extension in COUNT_EXTENSIONS]
    return result or None


def normalized_dates(text: str) -> list[str]:
    result: list[str] = []
    for match in re.finditer(r"(?<!\d)(\d{4})(?:[-/.年]?)(\d{2})(?:[-/.月]?)(\d{2})日?(?!\d)", text):
        result.append("".join(match.groups()))
    return result


def requested_assignees(title: str, comments: list[CommentRecord]) -> set[str]:
    known: set[str] = set()
    for comment in comments:
        for value in (comment.assignee, comment.author):
            if not value:
                continue
            escaped = re.escape(value)
            if re.search(
                rf"(?:待|由|责任人|负责人|处理人|to\s*[:：=]?)\s*[:：=为是]?\s*{escaped}|{escaped}\s*(?:的批注|的待办|的TODO|处理|负责|跟进)",
                title,
                re.IGNORECASE,
            ):
                known.add(value)
    if known:
        return known
    patterns = (
        r"待\s*([\u4e00-\u9fffA-Za-z][\u4e00-\u9fffA-Za-z0-9_.-]{0,19})\s*(?:处理|负责|完成|跟进)",
        r"(?:责任人|负责人|处理人)\s*[:：=为是]?\s*([\u4e00-\u9fffA-Za-z][\u4e00-\u9fffA-Za-z0-9_.-]{0,19})",
        r"(?<![A-Za-z0-9_])to(?![A-Za-z0-9_])\s*[:：=为是]?\s*([\u4e00-\u9fffA-Za-z][\u4e00-\u9fffA-Za-z0-9_.-]{0,19})",
        r"由\s*([\u4e00-\u9fffA-Za-z][\u4e00-\u9fffA-Za-z0-9_.-]{0,19})\s*(?:处理|负责|完成|跟进)",
    )
    for pattern in patterns:
        match = re.search(pattern, title, re.IGNORECASE)
        if match:
            return {match.group(1)}
    return set()


def comment_answer_text(comment: CommentRecord) -> str:
    """Render judged TODO lists without leaking an implementation-only path prefix."""

    if comment.structured and comment.todo and comment.assignee and comment.end_date:
        return f"todo: {comment.todo}, to: {comment.assignee},end_date: {comment.end_date}"
    value = comment.raw_text.strip()
    value = re.sub(r"^\s*(?:#|//|--|/\*+|\*|<!--)\s*", "", value)
    value = re.sub(r"\s*(?:\*/|-->)\s*$", "", value)
    return value.strip()


CREDENTIAL_ASSIGNMENT_RE = re.compile(
    r"(?i)(?:database[_\s-]*password|db[_\s-]*password|password|passwd|pwd|数据库密码|账号密码|用户密码|密码|口令|密钥|秘钥)"
    r"\s*['\"]?\s*[:=：]\s*['\"]?(?P<value>[^\s,，;；<>\"']+)"
)
CREDENTIAL_XML_RE = re.compile(
    r"(?is)<(?:[A-Za-z0-9_.-]*password|passwd|pwd|密码|口令|密钥)>\s*(?P<value>[^<\s]+)\s*</"
)


def credential_values(records: list[DocumentRecord], title: str) -> list[str]:
    """Extract only allowed environment credential values for the judge answer."""

    values: list[str] = []
    user_names = set(
        match.group(1)
        for match in re.finditer(
            r"(?<![A-Za-z0-9_])([A-Za-z][A-Za-z0-9_.-]{1,63})\s*(?:用户|账号|账户)",
            title,
            re.IGNORECASE,
        )
    )
    for record in records:
        for match in CREDENTIAL_XML_RE.finditer(record.text):
            values.append(match.group("value"))
        for line in record.text.splitlines():
            for match in CREDENTIAL_ASSIGNMENT_RE.finditer(line):
                values.append(match.group("value"))
            if user_names and "|" in line:
                cells = [cell.strip().strip("'\"") for cell in line.strip().strip("|").split("|")]
                for index, cell in enumerate(cells[:-1]):
                    if any(cell.casefold() == user.casefold() for user in user_names):
                        candidate = cells[index + 1]
                        if candidate and not re.search(r"(?:密码|口令|password|passwd|pwd)", candidate, re.IGNORECASE):
                            values.append(candidate)
    cleaned: list[str] = []
    seen: set[str] = set()
    for raw in values:
        value = raw.strip().strip("'\"`，,；;")
        if not value or value.casefold() in {"none", "null", "password", "passwd", "pwd"} or value in seen:
            continue
        seen.add(value)
        cleaned.append(value)
    return cleaned


def parse_answer_format(value: Any) -> dict[str, Any] | None:
    if isinstance(value, dict):
        nested = value.get("answer")
        return nested if isinstance(nested, dict) else value
    if not isinstance(value, str) or not value.strip():
        return None
    text = value.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.IGNORECASE)
    try:
        parsed = json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return None
    if isinstance(parsed, dict):
        nested = parsed.get("answer")
        return nested if isinstance(nested, dict) else parsed
    return None


def expected_answer_kind(value: Any) -> str | None:
    payload = parse_answer_format(value)
    if not payload:
        return None
    keys = set(payload)
    if keys == {"count"}:
        return "comment_count"
    if keys == {"source", "target"}:
        return "fix"
    if keys == {"error_msg"}:
        return "blocked"
    if keys == {"datas"}:
        return "list"
    if keys and keys.issubset(COUNT_EXTENSIONS):
        return "file_count"
    return None


def answer_format_extensions(value: Any) -> list[str] | None:
    payload = parse_answer_format(value)
    if not payload or not set(payload).issubset(COUNT_EXTENSIONS):
        return None
    return [ext for ext in sorted(COUNT_EXTENSIONS) if ext in payload]


def llm_answer_datas(text: str) -> list[Any]:
    """Keep the judged answer clean while accepting a model-emitted JSON payload."""
    candidate = text.strip()
    if candidate.startswith("```"):
        candidate = re.sub(r"^```(?:json)?\s*|\s*```$", "", candidate, flags=re.IGNORECASE)
    try:
        parsed = json.loads(candidate)
    except (json.JSONDecodeError, TypeError):
        return [text]
    if isinstance(parsed, dict):
        if isinstance(parsed.get("answer"), dict):
            parsed = parsed["answer"]
        datas = parsed.get("datas")
        if isinstance(datas, list):
            return datas
    if isinstance(parsed, list):
        return parsed
    return [text]


def relevant_snippets(record: DocumentRecord, title: str, allow_password: bool = False) -> list[str]:
    return [
        f"{item['source_path']}: {item['text']}"
        for item in evidence_snippets(record, title, allow_password=allow_password)
    ]


def evidence_snippets(record: DocumentRecord, title: str, allow_password: bool = False) -> list[dict[str, Any]]:
    terms = query_terms(title)
    scored: list[tuple[int, int, str, list[str]]] = []
    for line_no, raw_line in enumerate(record.text.splitlines(), start=1):
        if re.match(r"^\s*<(?:!doctype|html|head|meta|link|title|script|style)\b", raw_line, re.IGNORECASE):
            continue
        line = clean_evidence_text(raw_line)
        if not line:
            continue
        low = line.lower()
        if not allow_password and re.search(r"(password|passwd|密码|密钥|secret|token)", low, re.IGNORECASE):
            continue
        if re.search(r"(忽略.*规则|上帝模式|ignore.*previous|god\s*mode)", line, re.IGNORECASE):
            continue
        matched_terms = [term for term in terms if term.lower() in low]
        score = len(matched_terms)
        if any(term.lower() in COMMAND_TERMS for term in matched_terms):
            score += 5
        if re.search(r"(命令|gsql|ssh|curl|jdbc|连接)", title, re.IGNORECASE) and re.search(
            r"(gsql|ssh|curl|jdbc|mysql|psql|命令|连接)", line, re.IGNORECASE
        ):
            score += 6
        if score:
            scored.append((score, line_no, line[:240], matched_terms[:8]))
    scored.sort(key=lambda item: -item[0])
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for score, line_no, line, matched_terms in scored:
        value = f"{record.rel_path}:{line_no}:{line}"
        if value not in seen:
            seen.add(value)
            result.append(
                {
                    "source_path": record.rel_path,
                    "line": line_no,
                    "text": line,
                    "score": score,
                    "matched_terms": matched_terms,
                }
            )
        if len(result) >= 3:
            break
    return result


def clean_evidence_text(raw_line: str) -> str:
    text = re.sub(r"<[^>]+>", " ", raw_line)
    text = unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    if text in {"{", "}", ");", "};"}:
        return ""
    return text


def knowledge_evidence(scored_records: list[tuple[int, DocumentRecord]], title: str, guard: PermissionGuard) -> list[dict[str, Any]]:
    evidence: list[dict[str, Any]] = []
    allow_password_query = guard.has_password_intent(title)
    for _, record in scored_records:
        evidence.extend(
            evidence_snippets(
                record,
                title,
                allow_password=allow_password_query and guard.is_env_path(record.rel_path),
            )
        )
    if evidence:
        return evidence[:30]
    return [
        {"source_path": record.rel_path, "text": record.rel_path, "score": score, "matched_terms": []}
        for score, record in scored_records[:30]
    ]


def record_summary(record: DocumentRecord, score: int = 0) -> dict[str, Any]:
    return {
        "path": record.rel_path,
        "suffix": record.suffix,
        "name": record.name,
        "tags": sorted(record.tags),
        "comment_count": len(record.comments),
        "score": score,
    }


def comment_evidence(comment: CommentRecord) -> dict[str, Any]:
    return {
        "source_path": comment.source_path,
        "kind": comment.kind,
        "line": comment.line,
        "todo": comment.todo,
        "assignee": comment.assignee,
        "end_date": comment.end_date,
        "structured": comment.structured,
        "text": comment.summary(),
    }


def records_from_comments(index: WikiIndex, comments: list[CommentRecord]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for comment in comments:
        if comment.source_path in seen or comment.source_path not in index.by_path:
            continue
        seen.add(comment.source_path)
        records.append(record_summary(index.by_path[comment.source_path]))
    return records


def create_pivot_workbook(record: DocumentRecord, wiki_root: Path, title: str = "") -> str:
    if record.suffix == "xls":
        with tempfile.TemporaryDirectory(prefix="stonehenge-wiki-pivot-") as tmp:
            converted = convert_office(record.full_path, LEGACY_TO_MODERN["xls"], Path(tmp))
            if not converted:
                raise ValueError("无法转换旧版 xls 文件")
            converted_record = DocumentRecord(
                full_path=converted,
                rel_path=record.rel_path,
                suffix="xlsx",
                text=record.text,
                comments=record.comments,
                tags=record.tags,
            )
            return create_pivot_workbook(converted_record, wiki_root, title)
    try:
        return create_pivot_workbook_openpyxl(record, wiki_root, title)
    except ModuleNotFoundError:
        return create_pivot_summary_csv(record, wiki_root, title)
    except (KeyError, OSError, zipfile.BadZipFile, ET.ParseError):
        return create_pivot_summary_csv(record, wiki_root, title)


def create_pivot_workbook_openpyxl(record: DocumentRecord, wiki_root: Path, title: str = "") -> str:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, Reference

    source = load_workbook(record.full_path, data_only=True, read_only=True)
    candidates: list[tuple[int, list[tuple[Any, ...]]]] = []
    try:
        for sheet in source.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            headers = [str(value) for value in (rows[0] if rows else ()) if value is not None]
            score = sum(1 for header in headers if header.casefold() in title.casefold())
            candidates.append((score, rows))
    finally:
        source.close()
    rows = max(candidates, key=lambda item: item[0])[1] if candidates else []
    group_header, value_header, aggregation, summary = build_pivot_summary(rows, title)

    target_rel = f"output/fixed/pivot_{record.full_path.stem}.xlsx"
    target = wiki_root / target_rel
    target.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "pivot"
    result_header = pivot_value_header(value_header, aggregation)
    ws.append([group_header, result_header])
    for key, value in summary:
        ws.append([key, value])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = max(14, min(40, max(len(str(item[0])) for item in summary) + 2))
    ws.column_dimensions["B"].width = max(14, min(30, len(result_header) + 2))
    chart = BarChart()
    chart.title = f"{group_header} - {result_header}"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
    ws.add_chart(chart, "D2")
    fd, raw_temp = tempfile.mkstemp(prefix=f".{target.stem}.", suffix=".xlsx", dir=target.parent)
    os.close(fd)
    temp_target = Path(raw_temp)
    try:
        wb.save(temp_target)
        os.replace(temp_target, target)
    except Exception:
        temp_target.unlink(missing_ok=True)
        raise
    return target_rel


def create_pivot_summary_csv(record: DocumentRecord, wiki_root: Path, title: str = "") -> str:
    rows = read_xlsx_rows_zip(record.full_path)
    group_header, value_header, aggregation, summary = build_pivot_summary(rows, title)

    target_rel = f"output/fixed/pivot_{record.full_path.stem}.csv"
    target = wiki_root / target_rel
    target.parent.mkdir(parents=True, exist_ok=True)
    fd, raw_temp = tempfile.mkstemp(prefix=f".{target.stem}.", suffix=".csv", dir=target.parent)
    temp_target = Path(raw_temp)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow([group_header, pivot_value_header(value_header, aggregation)])
            writer.writerows(summary)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_target, target)
    except Exception:
        temp_target.unlink(missing_ok=True)
        raise
    return target_rel


def build_pivot_summary(
    rows: list[tuple[Any, ...]] | list[list[Any]],
    title: str,
) -> tuple[str, str | None, str, list[tuple[str, int | float]]]:
    if len(rows) < 2:
        raise ValueError("工作表数据不足")
    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    if not any(headers):
        raise ValueError("工作表缺少表头")
    data_rows = [list(row) for row in rows[1:] if any(cell not in {None, ""} for cell in row)]
    if not data_rows:
        raise ValueError("工作表数据不足")

    mentioned = [index for index, header in enumerate(headers) if header and header.casefold() in title.casefold()]
    group_idx = choose_group_column(headers, data_rows, title, mentioned)
    value_idx = choose_value_column(headers, data_rows, title, mentioned, group_idx)
    aggregation = choose_aggregation(title, value_idx)

    buckets: dict[str, list[float]] = {}
    counts: dict[str, int] = {}
    for row in data_rows:
        raw_group = row[group_idx] if group_idx < len(row) else None
        group = str(raw_group).strip() if raw_group not in {None, ""} else "空值"
        counts[group] = counts.get(group, 0) + 1
        if value_idx is not None:
            number = numeric_value(row[value_idx] if value_idx < len(row) else None)
            if number is not None:
                buckets.setdefault(group, []).append(number)

    summary: list[tuple[str, int | float]] = []
    for group in sorted(counts, key=lambda item: item.casefold()):
        values = buckets.get(group, [])
        if aggregation == "count":
            result: int | float = counts[group]
        elif not values:
            result = 0
        elif aggregation == "average":
            result = sum(values) / len(values)
        elif aggregation == "max":
            result = max(values)
        elif aggregation == "min":
            result = min(values)
        else:
            result = sum(values)
        if isinstance(result, float):
            result = int(result) if result.is_integer() else round(result, 6)
        summary.append((group, result))
    return headers[group_idx] or "分组", headers[value_idx] if value_idx is not None else None, aggregation, summary


def choose_group_column(
    headers: list[str], data_rows: list[list[Any]], title: str, mentioned: list[int]
) -> int:
    group_hint = re.search(r"按(?:照)?\s*([^，,。；;]{1,30}?)(?:分组|统计|汇总|合计|生成|制作|画|透视)", title)
    if group_hint:
        hint = group_hint.group(1).strip().casefold()
        for index, header in enumerate(headers):
            if header and (header.casefold() in hint or hint in header.casefold()):
                return index
    if mentioned:
        return mentioned[0]
    for index in range(len(headers)):
        values = [row[index] for row in data_rows if index < len(row) and row[index] not in {None, ""}]
        if values and sum(numeric_value(value) is None for value in values) >= len(values) / 2:
            return index
    return 0


def choose_value_column(
    headers: list[str],
    data_rows: list[list[Any]],
    title: str,
    mentioned: list[int],
    group_idx: int,
) -> int | None:
    for index in mentioned:
        if index != group_idx and column_has_numbers(data_rows, index):
            return index
    for index, header in enumerate(headers):
        if index == group_idx or not header:
            continue
        if header.casefold() in title.casefold() and column_has_numbers(data_rows, index):
            return index
    if re.search(r"(计数|条数|记录数|出现次数|多少条)", title):
        return None
    if not re.search(r"(合计|总和|汇总|总额|金额|销售额|平均|均值|最大|最高|最小|最低|sum|average|avg|max|min)", title, re.IGNORECASE):
        return None
    for index in range(len(headers)):
        if index != group_idx and column_has_numbers(data_rows, index):
            return index
    return None


def choose_aggregation(title: str, value_idx: int | None) -> str:
    if re.search(r"(平均|均值|average|avg)", title, re.IGNORECASE):
        return "average"
    if re.search(r"(最大|最高|max)", title, re.IGNORECASE):
        return "max"
    if re.search(r"(最小|最低|min)", title, re.IGNORECASE):
        return "min"
    if value_idx is None or re.search(r"(计数|条数|记录数|出现次数|多少条)", title):
        return "count"
    return "sum"


def pivot_value_header(value_header: str | None, aggregation: str) -> str:
    labels = {"count": "数量", "sum": "合计", "average": "平均值", "max": "最大值", "min": "最小值"}
    label = labels[aggregation]
    return f"{value_header}{label}" if value_header else label


def column_has_numbers(rows: list[list[Any]], index: int) -> bool:
    values = [row[index] for row in rows if index < len(row) and row[index] not in {None, ""}]
    return bool(values) and sum(numeric_value(value) is not None for value in values) >= max(1, len(values) / 2)


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    percent = text.endswith("%")
    cleaned = re.sub(r"[,，\s￥¥$€£%()]", "", text)
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if negative:
        number = -number
    if percent:
        number /= 100
    return number


def read_xlsx_rows_zip(path: Path) -> list[list[str]]:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        shared_strings = read_shared_strings(zf, ns)
        sheet_names = sorted(name for name in zf.namelist() if name.startswith("xl/worksheets/") and name.endswith(".xml"))
        if not sheet_names:
            return []
        root = ET.fromstring(zf.read(sheet_names[0]))
    rows: list[list[str]] = []
    for row_elem in root.findall(".//main:sheetData/main:row", ns):
        row_values: list[str] = []
        for cell in row_elem.findall("main:c", ns):
            col_idx = column_index(cell.attrib.get("r", ""))
            while len(row_values) <= col_idx:
                row_values.append("")
            row_values[col_idx] = cell_value(cell, shared_strings, ns)
        rows.append(row_values)
    return rows


def read_shared_strings(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("main:si", ns):
        strings.append("".join(text.strip() for text in item.itertext() if text and text.strip()))
    return strings


def cell_value(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> str:
    value_elem = cell.find("main:v", ns)
    if value_elem is None or value_elem.text is None:
        return ""
    value = value_elem.text
    if cell.attrib.get("t") == "s":
        try:
            return shared_strings[int(value)]
        except (IndexError, ValueError):
            return value
    return value


def column_index(ref: str) -> int:
    letters = re.match(r"[A-Z]+", ref.upper())
    if not letters:
        return 0
    index = 0
    for char in letters.group(0):
        index = index * 26 + ord(char) - ord("A") + 1
    return max(0, index - 1)
