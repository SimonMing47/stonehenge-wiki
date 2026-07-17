from __future__ import annotations

import re
import subprocess
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from .models import CommentRecord, DocumentRecord
from .office_bridge import LEGACY_TO_MODERN, convert_office

COUNT_EXTENSIONS = {
    "doc",
    "docx",
    "ppt",
    "pptx",
    "xls",
    "xlsx",
    "xml",
    "java",
    "py",
    "html",
    "md",
    "js",
}
SUPPORTED_EXTENSIONS = COUNT_EXTENSIONS | {"pdf"}

TEXT_EXTENSIONS = {"xml", "java", "py", "html", "md", "js", "txt", "csv", "json", "yaml", "yml"}
CODE_COMMENT_EXTENSIONS = {"xml", "java", "py", "html", "md", "js"}

STRUCTURED_TODO_RE = re.compile(
    r"(?is)\btodo\b\s*[:：]\s*(?P<todo>.*?)\s*(?:[,，;；]\s*|\s+)"
    r"\bto\b\s*[:：]\s*(?P<to>.*?)\s*(?:[,，;；]\s*|\s+)"
    r"\bend[\s_-]*date\b\s*[:：]\s*(?P<end_date>\d{8})"
)
BLOCK_COMMENT_RE = re.compile(r"/\*.*?\*/|<!--.*?-->", re.DOTALL)
STRUCTURED_FIELD_RE = re.compile(
    r"(?i)(?<![A-Za-z0-9_])(?P<key>todo|to|end[\s_-]*date)\s*[:：]"
)
DATE_VALUE_RE = re.compile(
    r"(?<!\d)(?P<year>\d{4})(?:[-/.年]?)(?P<month>\d{2})(?:[-/.月]?)(?P<day>\d{2})日?(?!\d)"
)
ACTIONABLE_COMMENT_RE = re.compile(
    r"(?i)(?:\btodo\b|待(?:办|处理|实现|优化|修复|补充)?|需要|需(?:要)?|应该|"
    r"重构|优化|修复|调整|补充|有误|错误|改成|改为|替换|删除)"
)


def extract_document(path: Path, wiki_root: Path) -> DocumentRecord:
    suffix = path.suffix.lower().lstrip(".")
    rel_path = path.relative_to(wiki_root).as_posix()
    text, office_comments = extract_text_and_office_comments(path, suffix, rel_path)
    if suffix in CODE_COMMENT_EXTENSIONS:
        inline_comments = extract_inline_comments(rel_path, text, suffix)
    elif suffix in LEGACY_TO_MODERN and not office_comments:
        # Some legacy converters expose annotations as text paragraphs instead of
        # preserving an OOXML comment part. Only structured TODO metadata is
        # recognized by this fallback, preventing ordinary body prose from being
        # counted as a comment.
        fallback_comment = make_comment(rel_path, text, "office")
        inline_comments = [fallback_comment] if fallback_comment.structured else []
    else:
        inline_comments = []
    comments = office_comments + inline_comments
    return DocumentRecord(
        full_path=path,
        rel_path=rel_path,
        suffix=suffix,
        text=text,
        comments=dedupe_comments(comments),
        tags=infer_tags(rel_path, text),
    )


def extract_text_and_office_comments(
    path: Path, suffix: str, rel_path: str
) -> tuple[str, list[CommentRecord]]:
    # Some generators create a valid OOXML package but retain a legacy extension.
    # Sniffing the archive keeps those files usable even when LibreOffice is absent.
    archive_kind = detect_ooxml_kind(path)
    if archive_kind == "docx":
        return extract_docx(path, rel_path)
    if archive_kind == "pptx":
        return extract_pptx(path, rel_path)
    if archive_kind == "xlsx":
        return extract_xlsx(path, rel_path)
    if suffix == "docx":
        return extract_docx(path, rel_path)
    if suffix == "pptx":
        return extract_pptx(path, rel_path)
    if suffix == "xlsx":
        return extract_xlsx(path, rel_path)
    if suffix == "pdf":
        return extract_pdf(path), []
    if suffix in LEGACY_TO_MODERN:
        converted = extract_legacy_office(path, suffix, rel_path)
        if converted is not None:
            return converted
    if suffix in TEXT_EXTENSIONS:
        return read_text_best_effort(path), []
    if looks_like_text(path):
        return read_text_best_effort(path), []
    return extract_binary_strings(path), []


def detect_ooxml_kind(path: Path) -> str | None:
    """Return the OOXML family for a ZIP package without trusting its suffix."""
    try:
        if not zipfile.is_zipfile(path):
            return None
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
    except (OSError, zipfile.BadZipFile):
        return None
    if "word/document.xml" in names:
        return "docx"
    if "ppt/presentation.xml" in names or any(name.startswith("ppt/slides/slide") for name in names):
        return "pptx"
    if "xl/workbook.xml" in names:
        return "xlsx"
    return None


def extract_pdf(path: Path) -> str:
    pdftotext = shutil_which("pdftotext")
    if pdftotext:
        try:
            result = subprocess.run(
                [pdftotext, "-layout", str(path), "-"],
                text=True,
                capture_output=True,
                timeout=20,
                check=False,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout
        except Exception:
            pass
    return extract_binary_strings(path)


def shutil_which(name: str) -> str | None:
    from shutil import which

    return which(name)


def extract_legacy_office(path: Path, suffix: str, rel_path: str) -> tuple[str, list[CommentRecord]] | None:
    with tempfile.TemporaryDirectory(prefix="stonehenge-wiki-office-") as tmp:
        converted = convert_office(path, LEGACY_TO_MODERN[suffix], Path(tmp))
        if not converted:
            return None
        return extract_text_and_office_comments(converted, converted.suffix.lower().lstrip("."), rel_path)


def read_text_best_effort(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def looks_like_text(path: Path, sample_bytes: int = 65536) -> bool:
    try:
        with path.open("rb") as handle:
            data = handle.read(sample_bytes)
    except OSError:
        return False
    if not data:
        return True
    if b"\x00" in data:
        return False
    control = sum(byte < 9 or 13 < byte < 32 for byte in data)
    if control / len(data) > 0.03:
        return False
    for encoding in ("utf-8", "gb18030"):
        try:
            data.decode(encoding)
            return True
        except UnicodeDecodeError:
            continue
    return all(byte in b"\t\n\r" or 32 <= byte <= 126 for byte in data)


def extract_binary_strings(path: Path) -> str:
    data = path.read_bytes()
    chunks = re.findall(rb"[\x09\x0a\x0d\x20-\x7e]{4,}|(?:[\x80-\xff][\x80-\xff]?){4,}", data)
    decoded: list[str] = []
    for chunk in chunks:
        for encoding in ("utf-8", "gb18030", "latin-1"):
            try:
                decoded.append(chunk.decode(encoding))
                break
            except UnicodeDecodeError:
                continue
    return "\n".join(decoded)


def extract_docx(path: Path, rel_path: str) -> tuple[str, list[CommentRecord]]:
    text_parts: list[str] = []
    comments: list[CommentRecord] = []
    try:
        with zipfile.ZipFile(path) as zf:
            for name in zf.namelist():
                if name.startswith("word/") and name.endswith(".xml") and (
                    "document" in name or "header" in name or "footer" in name
                ):
                    text_parts.append(xml_text(zf.read(name)))
            if "word/comments.xml" in zf.namelist():
                comments.extend(parse_docx_comments(zf.read("word/comments.xml"), rel_path))
    except Exception:
        return extract_binary_strings(path), []
    return "\n".join(text_parts), comments


def parse_docx_comments(data: bytes, rel_path: str) -> list[CommentRecord]:
    comments: list[CommentRecord] = []
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return comments
    for elem in root.iter():
        if not elem.tag.endswith("}comment") and elem.tag != "comment":
            continue
        raw = " ".join(t.strip() for t in elem.itertext() if t and t.strip())
        if not raw:
            continue
        comments.append(
            make_comment(
                rel_path,
                raw,
                "office",
                author=attribute_by_local_name(elem, "author"),
                created=attribute_by_local_name(elem, "date"),
            )
        )
    return comments


def extract_pptx(path: Path, rel_path: str) -> tuple[str, list[CommentRecord]]:
    text_parts: list[str] = []
    comments: list[CommentRecord] = []
    try:
        with zipfile.ZipFile(path) as zf:
            authors = parse_pptx_authors(zf)
            for name in zf.namelist():
                if name.startswith("ppt/slides/") and name.endswith(".xml"):
                    text_parts.append(xml_text(zf.read(name)))
                if name.startswith("ppt/comments/") and name.endswith(".xml"):
                    comments.extend(parse_pptx_comments(zf.read(name), rel_path, authors))
    except Exception:
        return extract_binary_strings(path), []
    return "\n".join(text_parts), comments


def parse_pptx_authors(zf: zipfile.ZipFile) -> dict[str, str]:
    authors: dict[str, str] = {}
    candidates = [
        name
        for name in zf.namelist()
        if name in {"ppt/commentAuthors.xml", "ppt/authors.xml"}
        or (name.startswith("ppt/") and name.endswith("commentAuthors.xml"))
    ]
    for name in candidates:
        try:
            root = ET.fromstring(zf.read(name))
        except (ET.ParseError, KeyError):
            continue
        for elem in root.iter():
            if local_name(elem.tag) not in {"cmAuthor", "author"}:
                continue
            author_id = attribute_by_local_name(elem, "id")
            display_name = (
                attribute_by_local_name(elem, "name")
                or attribute_by_local_name(elem, "displayName")
                or " ".join(text.strip() for text in elem.itertext() if text and text.strip())
            )
            if author_id is not None and display_name:
                authors[str(author_id)] = display_name
    return authors


def parse_pptx_comments(data: bytes, rel_path: str, authors: dict[str, str]) -> list[CommentRecord]:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []
    result: list[CommentRecord] = []
    for elem in root.iter():
        if local_name(elem.tag) not in {"cm", "comment"}:
            continue
        text_values = [
            text.strip()
            for child in elem.iter()
            if local_name(child.tag) in {"text", "t"}
            for text in [child.text or ""]
            if text.strip()
        ]
        raw = " ".join(text_values) or " ".join(text.strip() for text in elem.itertext() if text and text.strip())
        if not raw:
            continue
        author_id = attribute_by_local_name(elem, "authorId")
        result.append(
            make_comment(
                rel_path,
                raw,
                "office",
                author=authors.get(str(author_id)) if author_id is not None else None,
                created=attribute_by_local_name(elem, "dt") or attribute_by_local_name(elem, "date"),
            )
        )
    return result


def extract_xlsx(path: Path, rel_path: str) -> tuple[str, list[CommentRecord]]:
    try:
        text, comments = extract_xlsx_openpyxl(path, rel_path)
        if not comments:
            _, zip_comments = extract_xlsx_zip(path, rel_path)
            comments = zip_comments
        return text, comments
    except Exception:
        return extract_xlsx_zip(path, rel_path)


def extract_xlsx_openpyxl(path: Path, rel_path: str) -> tuple[str, list[CommentRecord]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False, read_only=False)
    text_parts: list[str] = []
    comments: list[CommentRecord] = []
    for sheet in workbook.worksheets:
        text_parts.append(f"[sheet] {sheet.title}")
        for row in sheet.iter_rows():
            values: list[str] = []
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
                if cell.comment and cell.comment.text:
                    comments.append(
                        make_comment(
                            rel_path,
                            cell.comment.text,
                            "office",
                            author=cell.comment.author,
                        )
                    )
            if values:
                text_parts.append(" | ".join(values))
    return "\n".join(text_parts), comments


def extract_xlsx_zip(path: Path, rel_path: str) -> tuple[str, list[CommentRecord]]:
    text_parts: list[str] = []
    comments: list[CommentRecord] = []
    try:
        with zipfile.ZipFile(path) as zf:
            for name in zf.namelist():
                if name.endswith(".xml") and (
                    name.startswith("xl/worksheets/")
                    or name == "xl/sharedStrings.xml"
                    or name.startswith("xl/comments")
                    or name.startswith("xl/threadedComments/")
                ):
                    data = zf.read(name)
                    if name.startswith("xl/comments") or name.startswith("xl/threadedComments/"):
                        comments.extend(parse_xlsx_comments(data, rel_path))
                    else:
                        text_parts.append(xml_text(data))
    except Exception:
        return extract_binary_strings(path), []
    return "\n".join(text_parts), comments


def parse_xlsx_comments(data: bytes, rel_path: str) -> list[CommentRecord]:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return []
    authors = [
        " ".join(text.strip() for text in elem.itertext() if text and text.strip())
        for elem in root.iter()
        if local_name(elem.tag) == "author"
    ]
    result: list[CommentRecord] = []
    for elem in root.iter():
        if local_name(elem.tag) not in {"comment", "threadedComment"}:
            continue
        raw = " ".join(
            text.strip()
            for child in elem.iter()
            if local_name(child.tag) in {"t", "text"}
            for text in [child.text or ""]
            if text.strip()
        )
        if not raw:
            continue
        author: str | None = None
        author_id = attribute_by_local_name(elem, "authorId")
        if author_id is None:
            author_id = attribute_by_local_name(elem, "personId")
        if author_id is not None:
            try:
                author = authors[int(author_id)]
            except (ValueError, IndexError):
                pass
        result.append(
            make_comment(
                rel_path,
                raw,
                "office",
                author=author or author_id,
                created=attribute_by_local_name(elem, "dT") or attribute_by_local_name(elem, "date"),
            )
        )
    return result


def xml_text(data: bytes) -> str:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return data.decode("utf-8", errors="ignore")
    values: list[str] = []
    for elem in root.iter():
        if elem.text and elem.text.strip():
            values.append(elem.text.strip())
    return "\n".join(values)


def extract_inline_comments(rel_path: str, text: str, suffix: str, kind: str = "code") -> list[CommentRecord]:
    comments: list[CommentRecord] = []
    covered_lines: set[int] = set()
    for block in BLOCK_COMMENT_RE.finditer(text):
        start_line = text.count("\n", 0, block.start()) + 1
        end_line = start_line + block.group(0).count("\n")
        covered_lines.update(range(start_line, end_line + 1))
        raw = block.group(0)
        if parse_structured_fields(raw) or ACTIONABLE_COMMENT_RE.search(raw):
            comments.append(make_comment(rel_path, raw, kind, line=start_line))
    for line_no, line in enumerate(text.splitlines(), start=1):
        if line_no in covered_lines:
            continue
        structured = parse_structured_fields(line)
        if structured:
            comments.append(make_comment(rel_path, line, kind, line=line_no))
            continue
        if re.search(r"^\s*(#|//|--)", line) and ACTIONABLE_COMMENT_RE.search(line):
            comments.append(make_comment(rel_path, line, kind, line=line_no))
    return comments


def make_comment(
    rel_path: str,
    raw_text: str,
    kind: str,
    line: int | None = None,
    author: str | None = None,
    created: str | None = None,
) -> CommentRecord:
    cleaned = " ".join(raw_text.split())
    fields = parse_structured_fields(cleaned)
    if fields:
        return CommentRecord(
            source_path=rel_path,
            raw_text=cleaned,
            kind=kind,
            todo=fields["todo"],
            assignee=fields["to"],
            end_date=fields["end_date"],
            line=line,
            author=author,
            created=created,
            structured=True,
        )
    return CommentRecord(
        source_path=rel_path,
        raw_text=cleaned,
        kind=kind,
        line=line,
        author=author,
        created=created,
        structured=False,
    )


def clean_field(value: str) -> str:
    return value.strip(" \t\r\n,，;；。.*/#<>!-\"‘’“”'")


def parse_structured_fields(raw_text: str) -> dict[str, str] | None:
    """Parse loose TODO metadata regardless of field order, case, spacing, or colon style."""
    matches = list(STRUCTURED_FIELD_RE.finditer(raw_text))
    if not matches:
        return None
    values: dict[str, str] = {}
    for index, match in enumerate(matches):
        raw_key = re.sub(r"[\s_-]+", "_", match.group("key").lower())
        key = "end_date" if raw_key == "end_date" else raw_key
        end = matches[index + 1].start() if index + 1 < len(matches) else len(raw_text)
        value = clean_field(raw_text[match.end() : end])
        if key not in values and value:
            values[key] = value
    if not {"todo", "to", "end_date"}.issubset(values):
        return None
    date_match = DATE_VALUE_RE.search(values["end_date"])
    if not date_match:
        return None
    return {
        "todo": clean_field(values["todo"]),
        "to": clean_field(values["to"]),
        "end_date": "".join(date_match.group(name) for name in ("year", "month", "day")),
    }


def dedupe_comments(comments: list[CommentRecord]) -> list[CommentRecord]:
    seen: set[tuple[str, str, int | None, str | None, str | None]] = set()
    result: list[CommentRecord] = []
    for comment in comments:
        if comment.kind == "office":
            result.append(comment)
            continue
        key = (comment.source_path, comment.raw_text, comment.line, comment.author, comment.created)
        if key in seen:
            continue
        seen.add(key)
        result.append(comment)
    return result


def infer_tags(rel_path: str, text: str) -> set[str]:
    low = (rel_path + "\n" + text[:20000]).lower()
    tags: set[str] = set()
    if any(term in low for term in ("password", "passwd", "密码", "账号", "ip地址", "jdbc", "端口", "host")):
        tags.add("环境信息")
    if any(term in low for term in ("select ", "gauss", "高斯", "gsql", "数据库", "sql")):
        tags.add("数据库")
    if any(term in low for term in ("需求", "prd", "产品", "规则", "设计")):
        tags.add("需求设计")
    if any(term in low for term in ("命令", "ssh ", "curl ", "gsql ", "kubectl", "docker")):
        tags.add("常用命令")
    if any(term in low for term in ("业务", "流程", "客户", "订单", "合同")):
        tags.add("业务")
    return tags


def _attr(name: str) -> str:
    return f"{{http://schemas.openxmlformats.org/wordprocessingml/2006/main}}{name}"


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def attribute_by_local_name(elem: ET.Element, name: str) -> str | None:
    for key, value in elem.attrib.items():
        if local_name(key).lower() == name.lower():
            return value
    return None
